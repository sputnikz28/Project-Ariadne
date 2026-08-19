"""Tests for dashboard/excel_export.py. Builds synthetic DashboardDataset
fixtures via dashboard_data.py's own real builders (build_dashboard_dataset)
— never touches HeroRegistry/LegendRegistry/historical datasets/disk,
except the single test that exercises export_to_excel(), which always
writes inside tempfile.mkdtemp(), never into the real project tree.
"""

import dataclasses
import tempfile
import unittest
from pathlib import Path

from core.services.dashboard_data import (
    CharacterRow,
    DrawRow,
    EconomyDrawRow,
    EconomySummary,
    HeroRow,
    HouseEntry,
    LegendRow,
    PrizeCategoryAggregate,
    PrizeCategoryRow,
    PrizeCategorySummary,
    build_dashboard_dataset,
)
from dashboard.excel_export import (
    _SHEET_ORDER,
    build_workbook,
    export_to_excel,
)


# ---------------------------------------------------------------------------
# Row fixtures — one representative record per row type, override any
# field via keyword to build edge cases (None, "impossible" values, etc.)
# ---------------------------------------------------------------------------

def make_hero_row(**overrides):
    base = dict(
        hero_id="HERO-2026-057-000149f4",
        dedup_hash="dedup-hash-1",
        entity_id="H-1",
        entity_name="Test Hero",
        race="Elfo",
        generation=1,
        provenance="legacy",
        draw_id="057/2026",
        draw_date="2026-07-17",
        official_numeros=(1, 2, 3, 4, 5),
        official_estrelas=(1, 2),
        predicted_numeros=(1, 2, 3, 4, 5),
        predicted_estrelas=(1, 2),
        matched_numbers_count=5,
        matched_stars_count=2,
        hero_category="4+1",
        hero_tier="TIER_3",
        registered_at=None,
    )
    base.update(overrides)
    return HeroRow(**base)


def make_legend_row(**overrides):
    base = dict(
        legend_id="LEGEND-abc123",
        source_prediction_id="spid-1",
        entity_id="H-1",
        entity_name="Test Legend",
        race="Bruxa",
        promotion_draw="062/2026",
        promotion_draw_date="2026-08-04",
        promotion_threshold=3,
        promotion_tier="LEGEND_TIER_4",
        criteria_version="v1",
        hero_count=3,
        qualified_draws=3,
        provenance="legacy",
    )
    base.update(overrides)
    return LegendRow(**base)


def make_character_row(**overrides):
    base = dict(entity_id="H-1", nome="Test Character", raca="Elfo", titulo="Título", metodo=None, faccao=None)
    base.update(overrides)
    return CharacterRow(**base)


def make_house_entry(**overrides):
    base = dict(casa="Casa de Teste", declared_by_races=("Elfo",), observed_in_population=True, source="both")
    base.update(overrides)
    return HouseEntry(**base)


def make_draw_row(**overrides):
    base = dict(
        numero_sorteio="001/2026", data="2026-01-02", dia_semana="sexta-feira",
        numeros=(1, 2, 3, 4, 5), estrelas=(1, 2), soma=15, gaps=(1, 1, 1, 1), fase_lua="Cheia",
    )
    base.update(overrides)
    return DrawRow(**base)


def make_economy_draw_row(**overrides):
    base = dict(
        numero_sorteio="001/2026", data="2026-01-02",
        receita_liquida_apostas_eur=1000.5, montante_para_premios_eur=500.25,
        percentagem_receita_para_premios=50.0, previsao_proximo_jackpot_eur=17000000,
        registos_portugal=100, combinacoes_registadas_portugal=200,
        apostas_registadas_portugal=300, combinacoes_por_registo=2.0,
        apostas_por_registo=3.0, receita_media_por_aposta_eur=3.33,
        houve_vencedor_1_premio_total=True, houve_vencedor_1_premio_portugal=False,
        total_vencedores_todas_categorias=10, total_vencedores_portugal_todas_categorias=1,
        dados_financeiros_disponiveis=True, categorias_premio_disponiveis=True,
    )
    base.update(overrides)
    return EconomyDrawRow(**base)


def make_economy_summary(**overrides):
    base = dict(
        moeda="EUR", ano=2026, sorteios_no_periodo=64, sorteios_com_dados_financeiros=15,
        sorteios_com_previsao_jackpot=15, sorteios_com_vencedor_1_premio_total=2,
        sorteios_com_vencedor_1_premio_portugal=0, receita_liquida_apostas_total_eur=15000.0,
        montante_para_premios_total_eur=7500.0, percentagem_receita_para_premios_media=50.0,
        previsao_jackpot_minimo_eur=17000000, previsao_jackpot_maximo_eur=130000000,
        receita_media_por_aposta_eur_media=3.33, total_vencedores_todas_categorias_soma=150,
        total_vencedores_portugal_todas_categorias_soma=15,
        percentagem_sorteios_com_dados_financeiros=23.4375,
        receita_media_por_sorteio_com_dados_eur=1000.0, montante_medio_para_premios_eur=500.0,
        jackpot_medio_previsto_eur=50000000.0,
        percentagem_sorteios_com_vencedor_1_premio_total=100.0,
        nota="Dados financeiros reais disponíveis para 15 de 64 sorteios.",
    )
    base.update(overrides)
    return EconomySummary(**base)


def make_prize_category_row(**overrides):
    base = dict(
        numero_sorteio="001/2026", data="2026-01-02", categoria=1, acertos="5 números + 2 estrelas",
        vencedores_portugal=0, vencedores_total=1, percentagem_portugal_no_total=0.0,
        categorias_disponiveis=True,
    )
    base.update(overrides)
    return PrizeCategoryRow(**base)


def make_prize_category_aggregate(**overrides):
    base = dict(
        categoria=1, acertos="5 números + 2 estrelas", sorteios_com_dados=15,
        vencedores_portugal_total=0, vencedores_total_total=15,
        percentagem_portugal_no_total_media=0.0,
    )
    base.update(overrides)
    return PrizeCategoryAggregate(**base)


def make_prize_category_summary(por_categoria=None, **overrides):
    base = dict(
        ano=2026, sorteios_no_periodo=64, sorteios_com_categorias_disponiveis=15,
        percentagem_sorteios_com_categorias_disponiveis=23.4375,
        nota="Categorias de prémios reais disponíveis para 15 de 64 sorteios.",
    )
    base.update(overrides)
    return PrizeCategorySummary(
        por_categoria=tuple(por_categoria) if por_categoria is not None else (make_prize_category_aggregate(),),
        **base,
    )


_UNSET = object()


def make_dataset(
    heroes=None, legends=None, key_base=None, characters=None, houses=None,
    economy_draws=None, economy_summary=_UNSET,
    prize_category_rows=None, prize_category_summary=_UNSET,
    gerado_em=None,
):
    return build_dashboard_dataset(
        heroes=heroes if heroes is not None else [make_hero_row()],
        legends=legends if legends is not None else [make_legend_row()],
        key_base=key_base if key_base is not None else [make_draw_row()],
        characters=characters if characters is not None else [make_character_row()],
        houses=houses if houses is not None else [make_house_entry()],
        gerado_em=gerado_em,
        economy_draws=economy_draws if economy_draws is not None else [make_economy_draw_row()],
        economy_summary=economy_summary if economy_summary is not _UNSET else make_economy_summary(),
        prize_category_rows=prize_category_rows if prize_category_rows is not None else [make_prize_category_row()],
        prize_category_summary=(
            prize_category_summary if prize_category_summary is not _UNSET else make_prize_category_summary()
        ),
    )


# ---------------------------------------------------------------------------
# Scan helpers (test-only) — locate a labelled row within a vertical
# ("Field", "Value") block without assuming an exact row number, so tests
# stay stable if a block gains/loses an unrelated field elsewhere.
# ---------------------------------------------------------------------------

def _value_for_label(ws, label, max_row=200):
    for row in range(1, max_row + 1):
        if ws.cell(row=row, column=1).value == label:
            return ws.cell(row=row, column=2).value
    raise AssertionError(f"label {label!r} not found in sheet {ws.title!r}")


def _header_row(ws, expected_headers, max_row=200):
    width = len(expected_headers)
    for row in range(1, max_row + 1):
        values = tuple(ws.cell(row=row, column=col).value for col in range(1, width + 1))
        if values == tuple(expected_headers):
            return row
    raise AssertionError(f"header row {expected_headers!r} not found in sheet {ws.title!r}")


def _data_rows(ws, header_row, width):
    rows = []
    row = header_row + 1
    while ws.cell(row=row, column=1).value is not None:
        rows.append(tuple(ws.cell(row=row, column=col).value for col in range(1, width + 1)))
        row += 1
    return rows


class TestWorkbookStructure(unittest.TestCase):
    def test_workbook_has_exactly_the_expected_sheets_in_order(self):
        wb = build_workbook(make_dataset())
        self.assertEqual(tuple(wb.sheetnames), _SHEET_ORDER)


class TestExecutiveSummarySheet(unittest.TestCase):
    def test_workbook_metadata_is_written_verbatim(self):
        wb = build_workbook(make_dataset(), project_version="V99", generated_at="2026-08-18T00:00:00+00:00")
        ws = wb["Executive Summary"]
        self.assertEqual(_value_for_label(ws, "workbook_project_version"), "V99")
        self.assertEqual(_value_for_label(ws, "workbook_generated_at"), "2026-08-18T00:00:00+00:00")
        self.assertIn("derived", _value_for_label(ws, "workbook_note"))

    def test_metadata_defaults_to_none_when_not_supplied(self):
        wb = build_workbook(make_dataset())
        ws = wb["Executive Summary"]
        self.assertIsNone(_value_for_label(ws, "workbook_project_version"))
        self.assertIsNone(_value_for_label(ws, "workbook_generated_at"))

    def test_executive_and_economia_fields_come_from_dataset_verbatim(self):
        dataset = make_dataset(
            heroes=[make_hero_row(), make_hero_row(hero_id="HERO-2026-058-1")],
            legends=[],
            gerado_em="2026-08-01T00:00:00+00:00",
        )
        wb = build_workbook(dataset)
        ws = wb["Executive Summary"]
        self.assertEqual(_value_for_label(ws, "total_heroes"), 2)
        self.assertEqual(_value_for_label(ws, "total_legends"), 0)
        self.assertEqual(_value_for_label(ws, "gerado_em"), "2026-08-01T00:00:00+00:00")
        self.assertEqual(_value_for_label(ws, "economia_fonte_financeira"), "não configurada")


class TestTabularSheets(unittest.TestCase):
    def test_heroes_sheet(self):
        heroes = [make_hero_row(), make_hero_row(hero_id="HERO-2026-058-2", registered_at="2026-08-01T00:00:00+00:00")]
        wb = build_workbook(make_dataset(heroes=heroes))
        ws = wb["Heroes"]
        from dashboard.excel_export import _HERO_HEADERS
        header_row = _header_row(ws, _HERO_HEADERS)
        rows = _data_rows(ws, header_row, len(_HERO_HEADERS))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0][0], "HERO-2026-057-000149f4")
        self.assertEqual(rows[0][9], "1, 2, 3, 4, 5")  # official_numeros tuple -> joined string
        self.assertEqual(rows[1][-1], "2026-08-01T00:00:00+00:00")

    def test_legends_sheet_empty_is_headers_only(self):
        wb = build_workbook(make_dataset(legends=[]))
        ws = wb["Legends"]
        from dashboard.excel_export import _LEGEND_HEADERS
        header_row = _header_row(ws, _LEGEND_HEADERS)
        self.assertEqual(_data_rows(ws, header_row, len(_LEGEND_HEADERS)), [])

    def test_characters_sheet_none_fields_stay_none(self):
        wb = build_workbook(make_dataset(characters=[make_character_row(titulo=None, metodo=None)]))
        ws = wb["Characters"]
        from dashboard.excel_export import _CHARACTER_HEADERS
        header_row = _header_row(ws, _CHARACTER_HEADERS)
        rows = _data_rows(ws, header_row, len(_CHARACTER_HEADERS))
        titulo_idx = _CHARACTER_HEADERS.index("titulo")
        self.assertIsNone(rows[0][titulo_idx])

    def test_houses_sheet_boolean_and_tuple_fields(self):
        wb = build_workbook(make_dataset(houses=[make_house_entry(
            declared_by_races=("Elfo", "Anão"), observed_in_population=False,
        )]))
        ws = wb["Houses"]
        from dashboard.excel_export import _HOUSE_HEADERS
        header_row = _header_row(ws, _HOUSE_HEADERS)
        rows = _data_rows(ws, header_row, len(_HOUSE_HEADERS))
        self.assertEqual(rows[0][_HOUSE_HEADERS.index("declared_by_races")], "Elfo, Anão")
        observed = rows[0][_HOUSE_HEADERS.index("observed_in_population")]
        self.assertIs(observed, False)  # real bool, never 0

    def test_key_base_sheet_none_fase_lua_stays_none(self):
        wb = build_workbook(make_dataset(key_base=[make_draw_row(fase_lua=None)]))
        ws = wb["Key Base"]
        from dashboard.excel_export import _KEY_BASE_HEADERS
        header_row = _header_row(ws, _KEY_BASE_HEADERS)
        rows = _data_rows(ws, header_row, len(_KEY_BASE_HEADERS))
        self.assertIsNone(rows[0][_KEY_BASE_HEADERS.index("fase_lua")])


class TestEconomySheet(unittest.TestCase):
    def test_summary_fields_come_from_economy_summary_verbatim(self):
        summary = make_economy_summary(sorteios_com_dados_financeiros=999)  # deliberately implausible
        wb = build_workbook(make_dataset(economy_summary=summary))
        ws = wb["Economy"]
        self.assertEqual(_value_for_label(ws, "sorteios_com_dados_financeiros"), 999)

    def test_none_monetary_fields_stay_none_not_zero(self):
        draw = make_economy_draw_row(receita_liquida_apostas_eur=None, montante_para_premios_eur=None)
        wb = build_workbook(make_dataset(economy_draws=[draw]))
        ws = wb["Economy"]
        from dashboard.excel_export import _ECONOMY_DRAW_HEADERS
        header_row = _header_row(ws, _ECONOMY_DRAW_HEADERS)
        rows = _data_rows(ws, header_row, len(_ECONOMY_DRAW_HEADERS))
        self.assertIsNone(rows[0][_ECONOMY_DRAW_HEADERS.index("receita_liquida_apostas_eur")])
        self.assertIsNone(rows[0][_ECONOMY_DRAW_HEADERS.index("montante_para_premios_eur")])

    def test_monetary_values_stay_numeric(self):
        draw = make_economy_draw_row(receita_liquida_apostas_eur=1234.5)
        wb = build_workbook(make_dataset(economy_draws=[draw]))
        ws = wb["Economy"]
        from dashboard.excel_export import _ECONOMY_DRAW_HEADERS
        header_row = _header_row(ws, _ECONOMY_DRAW_HEADERS)
        rows = _data_rows(ws, header_row, len(_ECONOMY_DRAW_HEADERS))
        value = rows[0][_ECONOMY_DRAW_HEADERS.index("receita_liquida_apostas_eur")]
        self.assertEqual(value, 1234.5)
        self.assertIsInstance(value, float)

    def test_boolean_fields_are_not_coerced(self):
        draw = make_economy_draw_row(houve_vencedor_1_premio_total=True, houve_vencedor_1_premio_portugal=False)
        wb = build_workbook(make_dataset(economy_draws=[draw]))
        ws = wb["Economy"]
        from dashboard.excel_export import _ECONOMY_DRAW_HEADERS
        header_row = _header_row(ws, _ECONOMY_DRAW_HEADERS)
        rows = _data_rows(ws, header_row, len(_ECONOMY_DRAW_HEADERS))
        self.assertIs(rows[0][_ECONOMY_DRAW_HEADERS.index("houve_vencedor_1_premio_total")], True)
        self.assertIs(rows[0][_ECONOMY_DRAW_HEADERS.index("houve_vencedor_1_premio_portugal")], False)

    def test_summary_absent_shows_not_available_note_and_draws_table_still_present(self):
        wb = build_workbook(make_dataset(economy_summary=None, economy_draws=[make_economy_draw_row()]))
        ws = wb["Economy"]
        self.assertEqual(_value_for_label(ws, "economy_summary"), "not available in this dataset")
        from dashboard.excel_export import _ECONOMY_DRAW_HEADERS
        header_row = _header_row(ws, _ECONOMY_DRAW_HEADERS)
        self.assertEqual(len(_data_rows(ws, header_row, len(_ECONOMY_DRAW_HEADERS))), 1)


class TestPrizeCategoriesSheet(unittest.TestCase):
    def test_summary_and_aggregate_and_rows_come_from_dataset_verbatim(self):
        aggregate = make_prize_category_aggregate(categoria=2, vencedores_total_total=42)
        summary = make_prize_category_summary(por_categoria=[aggregate])
        row = make_prize_category_row(categoria=2, vencedores_total=7)
        wb = build_workbook(make_dataset(prize_category_summary=summary, prize_category_rows=[row]))
        ws = wb["Prize Categories"]

        self.assertEqual(_value_for_label(ws, "sorteios_no_periodo"), 64)

        from dashboard.excel_export import _PRIZE_CATEGORY_AGGREGATE_HEADERS, _PRIZE_CATEGORY_ROW_HEADERS
        agg_header_row = _header_row(ws, _PRIZE_CATEGORY_AGGREGATE_HEADERS)
        agg_rows = _data_rows(ws, agg_header_row, len(_PRIZE_CATEGORY_AGGREGATE_HEADERS))
        self.assertEqual(agg_rows[0][_PRIZE_CATEGORY_AGGREGATE_HEADERS.index("vencedores_total_total")], 42)

        row_header_row = _header_row(ws, _PRIZE_CATEGORY_ROW_HEADERS)
        rows = _data_rows(ws, row_header_row, len(_PRIZE_CATEGORY_ROW_HEADERS))
        self.assertEqual(rows[0][_PRIZE_CATEGORY_ROW_HEADERS.index("vencedores_total")], 7)

    def test_summary_absent_shows_not_available_and_rows_table_still_present(self):
        row = make_prize_category_row()
        wb = build_workbook(make_dataset(prize_category_summary=None, prize_category_rows=[row]))
        ws = wb["Prize Categories"]
        self.assertEqual(_value_for_label(ws, "prize_category_summary"), "not available in this dataset")
        from dashboard.excel_export import _PRIZE_CATEGORY_ROW_HEADERS
        header_row = _header_row(ws, _PRIZE_CATEGORY_ROW_HEADERS)
        self.assertEqual(len(_data_rows(ws, header_row, len(_PRIZE_CATEGORY_ROW_HEADERS))), 1)


class TestDeterminismAndPurity(unittest.TestCase):
    def _all_cells(self, wb):
        cells = {}
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cells[(name, cell.row, cell.column)] = (cell.value, type(cell.value))
        return cells

    def test_build_workbook_is_semantically_deterministic_for_same_input(self):
        dataset = make_dataset()
        wb1 = build_workbook(dataset, project_version="V12", generated_at="2026-08-18")
        wb2 = build_workbook(dataset, project_version="V12", generated_at="2026-08-18")
        self.assertEqual(self._all_cells(wb1), self._all_cells(wb2))

    def test_metadata_is_the_only_diff_when_only_metadata_changes(self):
        dataset = make_dataset()
        wb1 = build_workbook(dataset, project_version="V12", generated_at="2026-08-18")
        wb2 = build_workbook(dataset, project_version="V13", generated_at="2026-08-19")
        cells1 = self._all_cells(wb1)
        cells2 = self._all_cells(wb2)
        differing = {k for k in cells1 if cells1[k] != cells2.get(k)}
        differing |= {k for k in cells2 if cells2[k] != cells1.get(k)}
        for sheet, row, col in differing:
            self.assertEqual(sheet, "Executive Summary")
        metadata_labels = {"workbook_project_version", "workbook_generated_at"}
        ws1 = wb1["Executive Summary"]
        touched_labels = {ws1.cell(row=row, column=1).value for _, row, _ in differing}
        self.assertTrue(touched_labels.issubset(metadata_labels))

    def test_build_workbook_does_not_mutate_dataset(self):
        dataset = make_dataset()
        before = dataclasses.replace(dataset)  # shallow copy — fields are all frozen/immutable already
        build_workbook(dataset, project_version="V12", generated_at="2026-08-18")
        self.assertEqual(dataset, before)


class TestEmptyAndAbsentInputs(unittest.TestCase):
    def test_fully_empty_dataset_produces_valid_workbook_with_headers_only(self):
        dataset = build_dashboard_dataset(
            heroes=[], legends=[], key_base=[], characters=[], houses=[],
        )
        wb = build_workbook(dataset)
        self.assertEqual(tuple(wb.sheetnames), _SHEET_ORDER)

        from dashboard.excel_export import _HERO_HEADERS
        ws = wb["Heroes"]
        header_row = _header_row(ws, _HERO_HEADERS)
        self.assertEqual(_data_rows(ws, header_row, len(_HERO_HEADERS)), [])

        ws = wb["Economy"]
        self.assertEqual(_value_for_label(ws, "economy_summary"), "not available in this dataset")

        ws = wb["Prize Categories"]
        self.assertEqual(_value_for_label(ws, "prize_category_summary"), "not available in this dataset")

    def test_economy_and_prize_categories_absent_on_otherwise_populated_dataset(self):
        dataset = make_dataset(
            economy_draws=[], economy_summary=None,
            prize_category_rows=[], prize_category_summary=None,
        )
        wb = build_workbook(dataset)
        ws = wb["Economy"]
        self.assertEqual(_value_for_label(ws, "economy_summary"), "not available in this dataset")
        ws = wb["Prize Categories"]
        self.assertEqual(_value_for_label(ws, "prize_category_summary"), "not available in this dataset")


class TestExportToExcel(unittest.TestCase):
    def test_export_writes_a_loadable_file_in_tempfile_only(self):
        import openpyxl

        tmp_dir = tempfile.mkdtemp()
        try:
            path = Path(tmp_dir) / "subdir" / "dashboard.xlsx"
            result = export_to_excel(make_dataset(), path, project_version="V12", generated_at="2026-08-18")
            self.assertEqual(result, path)
            self.assertTrue(path.exists())

            loaded = openpyxl.load_workbook(path)
            self.assertEqual(tuple(loaded.sheetnames), _SHEET_ORDER)
        finally:
            import shutil
            shutil.rmtree(tmp_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
