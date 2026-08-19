"""Dashboard Excel export — first output/visualisation layer over
core.services.dashboard_data.DashboardDataset (V12.3, see CLAUDE.md
Roadmap).

Pure pass-through: every cell value written here is read directly from
an already-built DashboardDataset (or one of its nested row/summary
dataclasses) — never recomputed, aggregated, sorted, or inferred. The
only transformation applied is display formatting for cell compatibility
(tuple -> comma-joined string); None/bool/int/float pass through
unchanged (openpyxl preserves bool as boolean and never coerces None to
0 — verified before writing this module).

project_version/generated_at are accepted verbatim from the caller and
never defaulted here from VERSION or a clock — same discipline
dashboard_data.py already applies to ExecutiveSummary.gerado_em.

Only two functions are public: build_workbook() (pure — builds and
returns an in-memory openpyxl.Workbook, touches no disk) and
export_to_excel() (the only function that writes a file). Everything
else is a private, per-sheet helper.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.services.dashboard_data import DashboardDataset

_NOT_AVAILABLE = "not available in this dataset"

_SHEET_ORDER: tuple[str, ...] = (
    "Executive Summary",
    "Heroes",
    "Legends",
    "Characters",
    "Houses",
    "Key Base",
    "Economy",
    "Prize Categories",
)

_HERO_HEADERS: tuple[str, ...] = (
    "hero_id", "dedup_hash", "entity_id", "entity_name", "race", "generation",
    "provenance", "draw_id", "draw_date", "official_numeros", "official_estrelas",
    "predicted_numeros", "predicted_estrelas", "matched_numbers_count",
    "matched_stars_count", "hero_category", "hero_tier", "registered_at",
)

_LEGEND_HEADERS: tuple[str, ...] = (
    "legend_id", "source_prediction_id", "entity_id", "entity_name", "race",
    "promotion_draw", "promotion_draw_date", "promotion_threshold",
    "promotion_tier", "criteria_version", "hero_count", "qualified_draws",
    "provenance",
)

_CHARACTER_HEADERS: tuple[str, ...] = ("entity_id", "nome", "raca", "titulo", "metodo", "faccao")

_HOUSE_HEADERS: tuple[str, ...] = ("casa", "declared_by_races", "observed_in_population", "source")

_KEY_BASE_HEADERS: tuple[str, ...] = (
    "numero_sorteio", "data", "dia_semana", "numeros", "estrelas", "soma", "gaps", "fase_lua",
)

_ECONOMY_DRAW_HEADERS: tuple[str, ...] = (
    "numero_sorteio", "data", "receita_liquida_apostas_eur",
    "montante_para_premios_eur", "percentagem_receita_para_premios",
    "previsao_proximo_jackpot_eur", "registos_portugal",
    "combinacoes_registadas_portugal", "apostas_registadas_portugal",
    "combinacoes_por_registo", "apostas_por_registo",
    "receita_media_por_aposta_eur", "houve_vencedor_1_premio_total",
    "houve_vencedor_1_premio_portugal", "total_vencedores_todas_categorias",
    "total_vencedores_portugal_todas_categorias",
    "dados_financeiros_disponiveis", "categorias_premio_disponiveis",
)

_ECONOMY_SUMMARY_FIELDS: tuple[str, ...] = (
    "moeda", "ano", "sorteios_no_periodo", "sorteios_com_dados_financeiros",
    "sorteios_com_previsao_jackpot", "sorteios_com_vencedor_1_premio_total",
    "sorteios_com_vencedor_1_premio_portugal", "receita_liquida_apostas_total_eur",
    "montante_para_premios_total_eur", "percentagem_receita_para_premios_media",
    "previsao_jackpot_minimo_eur", "previsao_jackpot_maximo_eur",
    "receita_media_por_aposta_eur_media", "total_vencedores_todas_categorias_soma",
    "total_vencedores_portugal_todas_categorias_soma",
    "percentagem_sorteios_com_dados_financeiros",
    "receita_media_por_sorteio_com_dados_eur", "montante_medio_para_premios_eur",
    "jackpot_medio_previsto_eur", "percentagem_sorteios_com_vencedor_1_premio_total", "nota",
)

_PRIZE_CATEGORY_ROW_HEADERS: tuple[str, ...] = (
    "numero_sorteio", "data", "categoria", "acertos", "vencedores_portugal",
    "vencedores_total", "percentagem_portugal_no_total", "categorias_disponiveis",
)

_PRIZE_CATEGORY_AGGREGATE_HEADERS: tuple[str, ...] = (
    "categoria", "acertos", "sorteios_com_dados", "vencedores_portugal_total",
    "vencedores_total_total", "percentagem_portugal_no_total_media",
)

_PRIZE_CATEGORY_SUMMARY_FIELDS: tuple[str, ...] = (
    "ano", "sorteios_no_periodo", "sorteios_com_categorias_disponiveis",
    "percentagem_sorteios_com_categorias_disponiveis", "nota",
)

_ECONOMIA_PLACEHOLDER_FIELDS: tuple[str, ...] = (
    "investimento", "premios", "saldo", "roi", "fonte_financeira", "nota",
)


def _cell_value(raw: Any) -> Any:
    if isinstance(raw, tuple):
        return ", ".join(str(item) for item in raw)
    return raw


def _write_row(ws: Worksheet, row: int, values: Sequence[Any]) -> None:
    for col, value in enumerate(values, start=1):
        ws.cell(row=row, column=col).value = _cell_value(value)


def _write_table(ws: Worksheet, start_row: int, headers: Sequence[str], records: Sequence[Any]) -> int:
    """Writes `headers` at start_row, then one row per record — each
    value read via getattr(record, header). Returns the next free row.
    Zero records is valid: headers are still written, no data rows.
    """
    _write_row(ws, start_row, headers)
    row = start_row + 1
    for record in records:
        _write_row(ws, row, [getattr(record, header) for header in headers])
        row += 1
    return row


def _write_vertical(ws: Worksheet, start_row: int, pairs: Sequence[tuple[str, Any]]) -> int:
    """Writes a ("Field", "Value") header at start_row, then one row per
    (label, value) pair. Returns the next free row.
    """
    _write_row(ws, start_row, ("Field", "Value"))
    row = start_row + 1
    for label, value in pairs:
        _write_row(ws, row, (label, value))
        row += 1
    return row


def _build_executive_summary_sheet(
    ws: Worksheet, dataset: DashboardDataset, project_version: str | None, generated_at: str | None,
) -> None:
    executive = dataset.executive
    economia = executive.economia

    row = _write_vertical(ws, 1, (
        ("workbook_project_version", project_version),
        ("workbook_generated_at", generated_at),
        ("workbook_note", "This workbook is derived from DashboardDataset; it is not a source of truth."),
    ))
    row += 1
    row = _write_vertical(ws, row, (
        ("total_heroes", executive.total_heroes),
        ("total_legends", executive.total_legends),
        ("taxa_sucesso", executive.taxa_sucesso),
        ("diversidade_media", executive.diversidade_media),
        ("convergencia_media", executive.convergencia_media),
        ("geracoes_analisadas", executive.geracoes_analisadas),
        ("gerado_em", executive.gerado_em),
    ))
    row += 1
    _write_vertical(ws, row, tuple(
        (f"economia_{field}", getattr(economia, field)) for field in _ECONOMIA_PLACEHOLDER_FIELDS
    ))


def _build_economy_sheet(ws: Worksheet, dataset: DashboardDataset) -> None:
    summary = dataset.economy_summary
    if summary is not None:
        pairs = tuple((field, getattr(summary, field)) for field in _ECONOMY_SUMMARY_FIELDS)
    else:
        pairs = (("economy_summary", _NOT_AVAILABLE),)
    row = _write_vertical(ws, 1, pairs)
    row += 1
    _write_table(ws, row, _ECONOMY_DRAW_HEADERS, dataset.economy_draws)


def _build_prize_categories_sheet(ws: Worksheet, dataset: DashboardDataset) -> None:
    summary = dataset.prize_category_summary
    if summary is not None:
        pairs = tuple((field, getattr(summary, field)) for field in _PRIZE_CATEGORY_SUMMARY_FIELDS)
        aggregates = summary.por_categoria
    else:
        pairs = (("prize_category_summary", _NOT_AVAILABLE),)
        aggregates = ()
    row = _write_vertical(ws, 1, pairs)
    row += 1
    row = _write_table(ws, row, _PRIZE_CATEGORY_AGGREGATE_HEADERS, aggregates)
    row += 1
    _write_table(ws, row, _PRIZE_CATEGORY_ROW_HEADERS, dataset.prize_category_rows)


def build_workbook(
    dataset: DashboardDataset,
    *,
    project_version: str | None = None,
    generated_at: str | None = None,
) -> Workbook:
    """Pure — no disk I/O. Builds and returns an in-memory Workbook with
    exactly the 8 sheets in _SHEET_ORDER. project_version/generated_at
    are written verbatim into the Executive Summary metadata block and
    never sourced from VERSION or a clock inside this function.
    """
    wb = Workbook()
    wb.active.title = _SHEET_ORDER[0]
    sheets = {_SHEET_ORDER[0]: wb.active}
    for name in _SHEET_ORDER[1:]:
        sheets[name] = wb.create_sheet(name)

    _build_executive_summary_sheet(sheets["Executive Summary"], dataset, project_version, generated_at)
    _write_table(sheets["Heroes"], 1, _HERO_HEADERS, dataset.heroes)
    _write_table(sheets["Legends"], 1, _LEGEND_HEADERS, dataset.legends)
    _write_table(sheets["Characters"], 1, _CHARACTER_HEADERS, dataset.characters)
    _write_table(sheets["Houses"], 1, _HOUSE_HEADERS, dataset.houses)
    _write_table(sheets["Key Base"], 1, _KEY_BASE_HEADERS, dataset.key_base)
    _build_economy_sheet(sheets["Economy"], dataset)
    _build_prize_categories_sheet(sheets["Prize Categories"], dataset)
    return wb


def export_to_excel(
    dataset: DashboardDataset,
    path: str | Path,
    *,
    project_version: str | None = None,
    generated_at: str | None = None,
) -> Path:
    """The only function in this module that touches disk: builds the
    workbook via build_workbook() and saves it to `path`. Does not read
    Heroes, Legends, datasets, or any Registry — writes only the file the
    caller asked for.
    """
    wb = build_workbook(dataset, project_version=project_version, generated_at=generated_at)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
